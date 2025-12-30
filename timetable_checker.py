#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except Exception:  # pragma: no cover - optional for self-test
    openpyxl = None


DAYS_ORDER = ["월", "화", "수", "목", "금"]


class ParseError(Exception):
    pass


@dataclass
class DayBlock:
    day: str
    start_col: int


def load_sheet(path: str):
    if openpyxl is None:
        raise ParseError("openpyxl을 불러올 수 없습니다. requirements.txt를 설치하세요.")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise ParseError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    if "주간시간표" in wb.sheetnames:
        return wb["주간시간표"]
    if wb.sheetnames:
        return wb[wb.sheetnames[0]]
    raise ParseError("시트가 존재하지 않습니다.")


def detect_day_blocks(sheet) -> List[DayBlock]:
    day_blocks: List[DayBlock] = []
    seen = set()
    row = 2
    max_col = getattr(sheet, "max_column", 0) or 0
    for col in range(1, max_col + 1):
        value = sheet.cell(row=row, column=col).value
        if isinstance(value, str):
            value = value.strip()
        if value in DAYS_ORDER and value not in seen:
            day_blocks.append(DayBlock(day=value, start_col=col))
            seen.add(value)

    if not day_blocks:
        raise ParseError("요일 블록을 찾을 수 없습니다.")

    return day_blocks


def normalize_teacher_name(raw_name: str) -> str:
    name = raw_name.strip()
    name = re.sub(r"\s*\([^)]*\)\s*$", "", name)
    return name.strip()


def normalize_cell_text(value: str) -> str:
    text = value.replace("_x000D_\n", "\n")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text


def parse_cell_to_class(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    text = normalize_cell_text(value)
    first_line = text.split("\n", 1)[0]
    match = re.match(r"^\s*(\d+)", first_line)
    if not match:
        return None
    return match.group(1)


def format_class_code(code: str) -> str:
    if len(code) == 3 and code.isdigit():
        grade = code[0]
        class_no = int(code[1:])
        return f"{grade}학년 {class_no}반({code})"
    return f"({code})"


def parse_teacher_rows(sheet, day_blocks: List[DayBlock]) -> Dict[str, Dict[str, Dict[int, Optional[str]]]]:
    data: Dict[str, Dict[str, Dict[int, Optional[str]]]] = {}
    max_row = getattr(sheet, "max_row", 0) or 0
    for row in range(4, max_row + 1):
        teacher_cell = sheet.cell(row=row, column=1).value
        if teacher_cell is None:
            continue
        if isinstance(teacher_cell, str) and not teacher_cell.strip():
            continue
        teacher_name = normalize_teacher_name(str(teacher_cell))
        if not teacher_name:
            continue

        if teacher_name not in data:
            data[teacher_name] = {day: {p: None for p in range(1, 8)} for day in DAYS_ORDER}

        for block in day_blocks:
            for period in range(1, 8):
                col = block.start_col + period - 1
                value = sheet.cell(row=row, column=col).value
                class_code = parse_cell_to_class(value)
                data[teacher_name][block.day][period] = class_code

    return data


def analyze_patterns(
    data: Dict[str, Dict[str, Dict[int, Optional[str]]]],
    consecutive_len: int,
    target_periods: List[int],
    min_days: int,
    check_period7: bool,
) -> Tuple[Dict[str, List[str]], Dict]:
    messages: Dict[str, List[str]] = {}
    summary: Dict[str, Dict] = {}

    for teacher, day_map in data.items():
        teacher_msgs: List[str] = []
        teacher_summary = {
            "patternA": [],
            "patternB": {"triggered": False, "days": []},
            "patternC": {"triggered": False, "days": []},
        }

        # Pattern A
        seen_a = set()
        for day in DAYS_ORDER:
            periods = day_map.get(day, {})
            max_start = 7 - consecutive_len + 1
            for start in range(1, max_start + 1):
                window = [periods.get(p) for p in range(start, start + consecutive_len)]
                if any(code is None for code in window):
                    continue
                if len(set(window)) == 1:
                    class_code = window[0]
                    key = (day, start, start + consecutive_len - 1, class_code)
                    if key in seen_a:
                        continue
                    seen_a.add(key)
                    class_text = format_class_code(class_code)
                    teacher_msgs.append(
                        f"이 시간표는 {teacher} 선생님이 {day}요일에 {start}~{start + consecutive_len - 1}교시 연속 {class_text}입니다."
                    )
                    teacher_summary["patternA"].append(
                        {
                            "day": day,
                            "start": start,
                            "end": start + consecutive_len - 1,
                            "class_code": class_code,
                        }
                    )

        # Pattern B
        matched_days = []
        for day in DAYS_ORDER:
            periods = day_map.get(day, {})
            if all(periods.get(p) is not None for p in target_periods):
                matched_days.append(day)
        if len(matched_days) >= min_days:
            days_text = ",".join(matched_days)
            periods_text = ",".join(str(p) for p in target_periods)
            teacher_msgs.append(
                f"이 시간표는 {teacher} 선생님이 {len(DAYS_ORDER)}일 중 {min_days}일 이상 ({periods_text})교시에 수업이 있는 시간표입니다. (해당 요일: {days_text})"
            )
            teacher_summary["patternB"] = {"triggered": True, "days": matched_days}

        # Pattern C
        if check_period7:
            matched_days = []
            for day in DAYS_ORDER:
                periods = day_map.get(day, {})
                if periods.get(7) is not None:
                    matched_days.append(day)
            if len(matched_days) >= min_days:
                days_text = ",".join(matched_days)
                teacher_msgs.append(
                    f"이 시간표는 {teacher} 선생님이 {len(DAYS_ORDER)}일 중 {min_days}일 이상 7교시에 수업이 있는 시간표입니다. (해당 요일: {days_text})"
                )
                teacher_summary["patternC"] = {"triggered": True, "days": matched_days}

        if teacher_msgs:
            messages[teacher] = teacher_msgs
        summary[teacher] = teacher_summary

    return messages, summary


def format_report(messages: Dict[str, List[str]]) -> str:
    if not messages:
        return "문제 패턴이 발견되지 않았습니다."

    lines: List[str] = []
    for teacher in sorted(messages.keys()):
        lines.append(f"=== {teacher} 선생님 ===")
        for msg in messages[teacher]:
            lines.append(f"- {msg}")
    return "\n".join(lines)


def write_text_output(text: str, output_path: Optional[str]):
    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text)
    else:
        print(text)


def write_json_output(summary: Dict, json_path: Optional[str]):
    payload = {"teachers": summary}
    if json_path:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    else:
        print(json.dumps(payload, ensure_ascii=False, indent=2))


class FakeWorksheet:
    def __init__(self, data: List[List[Optional[str]]]):
        self._data = data
        self.max_row = len(data)
        self.max_column = max((len(row) for row in data), default=0)

    def cell(self, row: int, column: int):
        class Cell:
            def __init__(self, value):
                self.value = value

        try:
            value = self._data[row - 1][column - 1]
        except IndexError:
            value = None
        return Cell(value)


def build_sample_sheet_for_tests() -> FakeWorksheet:
    data = [[None] * 36 for _ in range(6)]
    data[0][0] = "제목"
    # Row 2: day headers
    for idx, day in enumerate(DAYS_ORDER):
        data[1][1 + idx * 7] = day
    # Row 3: period numbers
    for idx in range(len(DAYS_ORDER)):
        for p in range(7):
            data[2][1 + idx * 7 + p] = p + 1
    # Row 4: teacher name
    data[3][0] = "홍길동(1)"
    # Pattern A: 수요일 1~4교시 101
    wed_start = 1 + DAYS_ORDER.index("수") * 7
    for p in range(4):
        data[3][wed_start + p] = "101\n국어"
    # Pattern B: 월수금 1,4,5,7
    for day in ["월", "수", "금"]:
        start = 1 + DAYS_ORDER.index(day) * 7
        for p in [1, 4, 5, 7]:
            data[3][start + p - 1] = "203\n수학"
    # Pattern C: 화목금 7교시
    for day in ["화", "목", "금"]:
        start = 1 + DAYS_ORDER.index(day) * 7
        data[3][start + 6] = "305\n영어"
    return FakeWorksheet(data)


def run_self_tests():
    sheet = build_sample_sheet_for_tests()
    day_blocks = detect_day_blocks(sheet)
    data = parse_teacher_rows(sheet, day_blocks)
    messages, summary = analyze_patterns(
        data,
        consecutive_len=4,
        target_periods=[1, 4, 5, 7],
        min_days=3,
        check_period7=True,
    )
    report = format_report(messages)
    assert "홍길동" in report
    assert "수요일에 1~4교시 연속" in report
    assert "(1,4,5,7)교시에" in report
    assert "7교시에 수업" in report
    assert summary["홍길동"]["patternA"], "Pattern A 요약 누락"


def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    def pick_file():
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            file_var.set(path)

    def parse_int_list(value: str) -> List[int]:
        parts = [p.strip() for p in value.split(",") if p.strip()]
        return [int(p) for p in parts]

    def run_check():
        file_path = file_var.get().strip()
        if not file_path:
            messagebox.showwarning("입력 필요", "엑셀 파일을 선택하세요.")
            return

        try:
            min_days = int(min_days_var.get())
            consecutive_len = int(consecutive_var.get())
            target_periods = parse_int_list(target_periods_var.get())
            check_period7 = bool(check_period7_var.get())
        except ValueError:
            messagebox.showerror("입력 오류", "숫자 입력값을 확인하세요.")
            return

        try:
            sheet = load_sheet(file_path)
            day_blocks = detect_day_blocks(sheet)
            data = parse_teacher_rows(sheet, day_blocks)
            messages, summary = analyze_patterns(
                data,
                consecutive_len=consecutive_len,
                target_periods=target_periods,
                min_days=min_days,
                check_period7=check_period7,
            )
            report = format_report(messages)
            output_text.delete("1.0", tk.END)
            output_text.insert(tk.END, report)

            if json_var.get():
                json_report = json.dumps({"teachers": summary}, ensure_ascii=False, indent=2)
                output_text.insert(tk.END, "\n\n[JSON 요약]\n")
                output_text.insert(tk.END, json_report)
        except ParseError as exc:
            messagebox.showerror("파싱 오류", f"파일 형식을 해석할 수 없습니다: {exc}")
        except Exception as exc:
            messagebox.showerror("실행 오류", f"파일 형식을 해석할 수 없습니다: {exc}")

    def save_report():
        content = output_text.get("1.0", tk.END).strip()
        if not content:
            messagebox.showwarning("저장 실패", "저장할 결과가 없습니다.")
            return
        path = filedialog.asksaveasfilename(
            title="결과 저장",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
        )
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        messagebox.showinfo("저장 완료", "결과가 저장되었습니다.")

    root = tk.Tk()
    root.title("교사 시간표 패턴 검사기 - made by 최세현")
    root.geometry("900x700")

    file_var = tk.StringVar()
    min_days_var = tk.StringVar(value="3")
    consecutive_var = tk.StringVar(value="4")
    target_periods_var = tk.StringVar(value="1,4,5,7")
    check_period7_var = tk.IntVar(value=1)
    json_var = tk.IntVar(value=0)

    top_frame = tk.Frame(root, padx=10, pady=10)
    top_frame.pack(fill=tk.X)

    tk.Label(top_frame, text="엑셀 파일").grid(row=0, column=0, sticky="w")
    tk.Entry(top_frame, textvariable=file_var, width=60).grid(row=0, column=1, padx=5)
    tk.Button(top_frame, text="찾아보기", command=pick_file).grid(row=0, column=2, padx=5)

    options_frame = tk.Frame(root, padx=10, pady=5)
    options_frame.pack(fill=tk.X)

    tk.Label(options_frame, text="최소 요일 수").grid(row=0, column=0, sticky="w")
    tk.Entry(options_frame, textvariable=min_days_var, width=6).grid(row=0, column=1, padx=5)

    tk.Label(options_frame, text="연속 교시 길이").grid(row=0, column=2, sticky="w")
    tk.Entry(options_frame, textvariable=consecutive_var, width=6).grid(row=0, column=3, padx=5)

    tk.Label(options_frame, text="대상 교시").grid(row=0, column=4, sticky="w")
    tk.Entry(options_frame, textvariable=target_periods_var, width=12).grid(row=0, column=5, padx=5)

    tk.Checkbutton(options_frame, text="7교시 검사", variable=check_period7_var).grid(row=0, column=6, padx=5)
    tk.Checkbutton(options_frame, text="JSON 요약 포함", variable=json_var).grid(row=0, column=7, padx=5)

    button_frame = tk.Frame(root, padx=10, pady=5)
    button_frame.pack(fill=tk.X)
    tk.Button(button_frame, text="검사 실행", command=run_check, width=12).pack(side=tk.LEFT)
    tk.Button(button_frame, text="결과 저장", command=save_report, width=12).pack(side=tk.LEFT, padx=5)

    output_text = tk.Text(root, wrap=tk.WORD)
    output_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    footer = tk.Label(root, text="made by 최세현", anchor="e")
    footer.pack(fill=tk.X, padx=10, pady=5)

    root.mainloop()


def main(argv: List[str]) -> int:
    if len(argv) >= 1 and argv[0] == "--self-test":
        run_self_tests()
        print("자체 테스트가 통과했습니다.")
        return 0
    run_gui()
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
